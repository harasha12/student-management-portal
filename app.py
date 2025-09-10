# app.py (UPDATED for MySQL Workbench)

import random
import html
import re
import os
from datetime import datetime, timedelta

from flask import Flask, Response, render_template, request, redirect, url_for, session, g
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
import mysql.connector
from mysql.connector import Error

from fpdf import FPDF
import PyPDF2
import os
secret_key = os.urandom(24)  # generates a random 24-byte key
print(secret_key)

print("TEMPLATE FOLDER PATH:", os.path.join(os.getcwd(), "templates"))

app = Flask(__name__)
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = 'harshavardhanvangara@gmail.com'       # replace with your email
app.config['MAIL_PASSWORD'] = 'bfllbazxsxinlheb'                   # replace with your app password
app.config['MAIL_DEFAULT_SENDER'] = 'harshavardhanvangara@gmail.com'
mail = Mail(app)
import binascii

key_bytes = b'$\xcc\xde\xdc\x1a8*%\xca\xeb~\xd2\x01RZa\xab6\xb9!\xfaI\xd4\x0c'
secret_key = binascii.hexlify(key_bytes).decode()  # convert to string

app.secret_key = b'$\xcc\xde\xdc\x1a8*%\xca\xeb~\xd2\x01RZa\xab6\xb9!\xfaI\xd4\x0c'
app.permanent_session_lifetime = timedelta(minutes=20)

# ---------------------
# MySQL connection
# ---------------------
from flask import g

def get_db():
    if 'db' not in g:
        g.db = mysql.connector.connect(
            host=os.getenv("DB_HOST", "metro.proxy.rlwy.net"),
            port=int(os.getenv("DB_PORT", 12600)),
            user=os.getenv("DB_USER", "root"),
            password=os.getenv("DB_PASS", "XfSjxukNhBBfCBJgIlXUTQVHazTrBfwc"),
            database=os.getenv("DB_NAME", "railway"),
            autocommit=False
        )
    return g.db

@app.teardown_appcontext
def close_db(error):
    db = g.pop('db', None)
    if db is not None and db.is_connected():
        db.close()

# ---------------------
# Utility to read PDF (unchanged)
# ---------------------
def read_pdf_text(filepath):
    text = ''
    with open(filepath, 'rb') as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            page_text = page.extract_text() or ''
            text += page_text
    return text
@app.route("/test_db")
def test_db():
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        cursor.execute("SELECT * FROM students LIMIT 5")
        rows = cursor.fetchall()
        cursor.close()
        return {"success": True, "data": rows}
    except Exception as e:
        return {"success": False, "error": str(e)}

# ---------------------
# Initialize DB (creates tables if not exist)
# ---------------------
def init_db():
    with app.app_context():
        db = get_db()
        cursor = db.cursor()
        try:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS students (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    student_id VARCHAR(100) UNIQUE,
                    username VARCHAR(100),
                    name VARCHAR(100),
                    email VARCHAR(100),
                    course VARCHAR(100),
                    enrollment_date DATE,
                    added_by VARCHAR(100)
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS attendance (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    student_id VARCHAR(100),
                    date DATE,
                    period1 VARCHAR(10),
                    period2 VARCHAR(10),
                    period3 VARCHAR(10),
                    period4 VARCHAR(10),
                    period5 VARCHAR(10),
                    period6 VARCHAR(10),
                    remarks TEXT,
                    added_by VARCHAR(100),
                    FOREIGN KEY(student_id) REFERENCES students(student_id)
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS grades (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    student_id VARCHAR(100),
                    subject VARCHAR(100),
                    grade VARCHAR(10),
                    FOREIGN KEY (student_id) REFERENCES students(student_id),
                    UNIQUE (student_id, subject)
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS messages (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    sender VARCHAR(100),
                    receiver VARCHAR(100),
                    message TEXT,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    username VARCHAR(100) UNIQUE,
                    email VARCHAR(255),
                    password VARCHAR(255),
                    role ENUM('teacher','student') NOT NULL DEFAULT 'student'
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS internal_marks (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    student_id VARCHAR(100) NOT NULL,
                    subject VARCHAR(100) NOT NULL,
                    marks VARCHAR(10) NOT NULL,
                    entered_by VARCHAR(100) NOT NULL
                )
            ''')
            db.commit()
        except Exception as e:
            db.rollback()
            print("init_db error:", e)
        finally:
            cursor.close()

# ---------------------
# Registration routes
# ---------------------
# Compatibility route: older templates might call url_for('register', role='teacher')

@app.route('/register/<role>', methods=['GET', 'POST'])
def register(role):
    role_lower = (role or '').strip().lower()
    if role_lower == 'student':
        return redirect(url_for('register_student'))
    elif role_lower == 'teacher':
        return redirect(url_for('register_teacher'))
    else:
        return render_template('register.html', error="Unknown role specified.")


def validate_password(password):
    """Validates password with complexity rules."""
    if len(password) < 6:
        return "Password must be at least 6 characters."
    if not re.search(r"\d", password):
        return "Password must include a number."
    if not re.search(r"[A-Z]", password):
        return "Password must include an uppercase letter."
    if not re.search(r"[!@#$%^&*()_+]", password):
        return "Password must include a special character."
    return None


@app.route('/register_student', methods=['GET', 'POST'])
def register_student():
    error, success = None, None
    username_from_url = request.args.get("username", "").strip()  # redirect case

    if request.method == 'POST':
        username = html.escape(request.form.get('username', username_from_url)).strip()
        email = html.escape(request.form.get('email', '')).strip()
        password = request.form.get('password', '')

        if not username or not password:
            error = "Please provide both username and password."
        else:
            error = validate_password(password)

        if not error:
            db = get_db()
            cursor = db.cursor(dictionary=True)
            try:
                hashed = generate_password_hash(password)

                # ✅ Only update existing student
                cursor.execute("""
                    UPDATE users
                    SET email = %s, password = %s, role = 'student'
                    WHERE username = %s AND role = 'student'
                """, (email, hashed, username))

                if cursor.rowcount == 0:
                    error = "⚠️ Student not found. Contact teacher."
                else:
                    db.commit()
                    flash("✅ Password set successfully! Please login.", "success")
                    return redirect(url_for('student_login'))
            finally:
                cursor.close()

    return render_template(
        'register.html',
        error=error,
        success=success,
        role='student',
        username=username_from_url
    )


@app.route('/register_teacher', methods=['GET', 'POST'])
def register_teacher():
    error, success = None, None
    if request.method == 'POST':
        username = html.escape(request.form.get('username', '')).strip()
        email = html.escape(request.form.get('email', '')).strip()
        password = request.form.get('password', '')

        if not username or not password:
            error = "Please provide both username and password."
        else:
            error = validate_password(password)

        if not error:
            db = get_db()
            cursor = db.cursor(dictionary=True)
            try:
                hashed = generate_password_hash(password)
                cursor.execute("""
                    INSERT INTO users (username, email, password, role)
                    VALUES (%s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE email = VALUES(email), password = VALUES(password), role = VALUES(role)
                """, (username, email, hashed, 'teacher'))  # ✅ Corrected role

                db.commit()
                success = "Teacher registered successfully!"
                return redirect(url_for('teacher_login'))
            except mysql.connector.IntegrityError:
                db.rollback()
                error = "Username already exists."
            except Exception as e:
                db.rollback()
                error = f"Error registering teacher: {e}"
            finally:
                cursor.close()

    return render_template('register.html', error=error, success=success, role='teacher')

# ---------------------
# Login routes
# ---------------------
@app.route('/teacher_login', methods=['GET', 'POST'])
def teacher_login():
    error = None

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        if not username or not password:
            error = "Please enter both username and password."
        else:
            db = get_db()
            cursor = db.cursor(dictionary=True, buffered=True)

            try:
                # Look up teacher in users table
                cursor.execute(
                    "SELECT * FROM users WHERE username = %s AND role = 'teacher' LIMIT 1",
                    (username,)
                )
                user = cursor.fetchone()

                if not user:
                    error = "No teacher found with that username."
                elif not check_password_hash(user['password'], password):
                    error = "Incorrect password."
                else:
                    session.permanent = True
                    session['username'] = user['username']
                    session['role'] = 'teacher'
                    return redirect(url_for('teacher_dashboard'))

            except Exception as e:
                error = f"Login error: {e}"
            finally:
                cursor.close()

    return render_template('teacher_login.html', error=error)


@app.route('/student_login', methods=['GET', 'POST'])
def student_login():
    error = None
    if request.method == 'POST':
        login_input = request.form.get('username', '').strip()  # can be Student ID or Email
        password = request.form.get('password', '')

        db = get_db()
        cursor = db.cursor(dictionary=True, buffered=True)
        try:
            # ✅ Find student by ID or Email
            cursor.execute(
                "SELECT * FROM users WHERE (username = %s OR email = %s) AND role = 'student'",
                (login_input, login_input)
            )
            user = cursor.fetchone()

            if not user:
                error = "⚠️ Student not found. Contact teacher."
            elif not user.get("password"):  
                # ✅ First-time login → redirect to register
                flash("⚠️ First-time login detected. Please register and set a password.", "warning")
                return redirect(url_for("register_student", username=user["username"]))
            elif check_password_hash(user["password"], password):
                # ✅ Success login
                session['user'] = user['username']
                session['role'] = 'student'
                return redirect(url_for('student_dashboard'))
            else:
                error = "❌ Invalid Student ID/Email or Password"
        except Exception as e:
            error = f"Error: {str(e)}"
        finally:
            cursor.close()

    return render_template('student_login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('choose_login'))

# ---------------------
# Dashboard
# ---------------------
# ---------------- TEACHER DASHBOARD ----------------
@app.route('/teacher_dashboard')
def teacher_dashboard():
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        username = session['username']  # teacher's username
        cursor.execute("SELECT * FROM users WHERE username = %s AND role = 'teacher'", (username,))
        teacher = cursor.fetchone()

        if not teacher:
            return "Unauthorized", 403

        # ✅ Fetch all students
        cursor.execute("SELECT student_id, name, course FROM students")
        students = cursor.fetchall()

        # ✅ Fetch all marks
        cursor.execute("SELECT * FROM internal_marks")
        marks = cursor.fetchall()

        # ✅ Fetch latest attendance records (optional)
        cursor.execute("SELECT * FROM attendance ORDER BY date DESC LIMIT 20")
        attendance = cursor.fetchall()

        return render_template('teacher_dashboard.html',
                               teacher=teacher,
                               students=students,
                               marks=marks,
                               attendance=attendance)
    finally:
        cursor.close()


# ---------------- STUDENT DASHBOARD ----------------
@app.route('/student_dashboard', methods=['GET', 'POST'])
def student_dashboard():
    if 'user' not in session or session.get('role') != 'student':
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor(dictionary=True)

    try:
        student_id = session['user']

        # ✅ Fetch student details
        cursor.execute("SELECT * FROM students WHERE student_id = %s", (student_id,))
        student = cursor.fetchone()
        if not student:
            return "Unauthorized or student not found", 403

        # ✅ Selected semester (default = 1)
        semester = request.args.get("semester", 1, type=int)

        # ✅ Internal marks for selected semester
        cursor.execute("""
            SELECT subject, marks, semester 
            FROM internal_marks 
            WHERE student_id = %s AND semester = %s
        """, (student_id, semester))
        marks = cursor.fetchall()

        # ✅ Fetch distinct semesters available for dropdown
        cursor.execute("""
            SELECT DISTINCT semester FROM internal_marks WHERE student_id = %s ORDER BY semester
        """, (student_id,))
        semesters = [row['semester'] for row in cursor.fetchall()]

        # ✅ Attendance (last 6 months)
        six_months_ago = (datetime.now() - timedelta(days=180)).strftime('%Y-%m-%d')
        cursor.execute("""
            SELECT date, status 
            FROM attendance
            WHERE student_id = %s AND date >= %s
            ORDER BY date DESC
        """, (student_id, six_months_ago))
        records = cursor.fetchall()

        total_present = sum(1 for r in records if r['status'] and r['status'].lower() == 'present')
        total_absent = sum(1 for r in records if r['status'] and r['status'].lower() == 'absent')
        total_days = len(records)
        attendance_percentage = (total_present / total_days * 100) if total_days > 0 else 0

        return render_template('student_dashboard.html',
                               student=student,
                               marks=marks,
                               semesters=semesters,
                               selected_semester=semester,
                               records=records,
                               total_present=total_present,
                               total_absent=total_absent,
                               attendance_percentage=round(attendance_percentage, 2))
    finally:
        cursor.close()

#--------------------
# Add / View / Edit / Delete Students
# ---------------------
from flask import flash

@app.route('/add_student', methods=['GET', 'POST'])
def add_student():
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect(url_for('choose_login'))

    if request.method == 'POST':
        student_id = request.form.get('student_id', '').strip().replace("_", "")
        username = request.form.get('username', '').strip()
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        course = request.form.get('course', '').strip()
        year = request.form.get('year', '').strip()
        date = request.form.get('date', '').strip()
        added_by = session.get('username') or 'unknown'

        if not student_id or not name or not email or not course or not year or not date:
            flash("⚠️ All fields are required.", "danger")
            return redirect(url_for('add_student'))

        db = get_db()
        cursor = db.cursor()
        try:
            # ✅ Insert / Update students table (no password here)
            cursor.execute(
                """
                INSERT INTO students (student_id, username, name, email, course, year, added_by, enrollment_date)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE 
                    name = VALUES(name),
                    email = VALUES(email),
                    course = VALUES(course),
                    year = VALUES(year),
                    added_by = VALUES(added_by),
                    enrollment_date = VALUES(enrollment_date)
                """,
                (student_id, username, name, email, course, year, added_by, date)
            )

            # ✅ Insert / Update users table for login (student registers later, so no password set here)
            cursor.execute(
                """
                INSERT INTO users (username, email, role)
                VALUES (%s, %s, 'student')
                ON DUPLICATE KEY UPDATE
                    email = VALUES(email),
                    role = VALUES(role)
                """,
                (student_id, email)
            )

            db.commit()
            flash("✅ Student added/updated successfully!", "success")
            return redirect(url_for('view_students'))

        except mysql.connector.IntegrityError:
            db.rollback()
            flash("⚠️ Student ID already exists.", "danger")
        except Exception as e:
            db.rollback()
            print("DEBUG error:", str(e))
            flash(f"❌ Error adding student: {str(e)}", "danger")
        finally:
            cursor.close()

    return render_template('add_student.html')




@app.route('/students')
def view_students():
    if 'username' not in session:
        return redirect('/')

    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM students")
        students = cursor.fetchall()
        return render_template('view_students.html', students=students)
    finally:
        cursor.close()

@app.route('/edit_student/<int:id>', methods=['GET', 'POST'])
def edit_student(id):
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        if request.method == 'POST':
            name = request.form.get('name')
            email = request.form.get('email')
            course = request.form.get('course')
            enrollment_date = request.form.get('enrollment_date')
            cur2 = db.cursor()
            try:
                cur2.execute(
                    "UPDATE students SET name=%s, email=%s, course=%s, enrollment_date=%s WHERE id=%s",
                    (name, email, course, enrollment_date, id)
                )
                db.commit()
                return redirect(url_for('view_students'))
            finally:
                cur2.close()

        cursor.execute("SELECT * FROM students WHERE id=%s", (id,))
        student = cursor.fetchone()
        return render_template('edit_student.html', student=student)
    finally:
        cursor.close()

@app.route('/delete_student/<int:id>')
def delete_student(id):
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute("SELECT student_id FROM students WHERE id = %s", (id,))
        res = cursor.fetchone()
        if res:
            student_id = res[0]
            cursor.execute("DELETE FROM attendance WHERE student_id = %s", (student_id,))
            cursor.execute("DELETE FROM internal_marks WHERE student_id = %s", (student_id,))
            cursor.execute("DELETE FROM students WHERE id = %s", (id,))
            db.commit()
    except Exception as e:
        db.rollback()
        print("delete_student error:", e)
    finally:
        cursor.close()

    return redirect(url_for('view_students'))

# ---------------------
# Attendance
# ---------------------
@app.route('/mark_attendance', methods=['GET', 'POST'])
def mark_attendance():
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor(dictionary=True)
    teacher_username = session['username']

    try:
        # ✅ Fetch ALL students (not only added_by, otherwise empty after Excel uploads)
        cursor.execute("SELECT student_id, name FROM students")
        students = cursor.fetchall()

        selected_period = None

        if request.method == 'POST':
            date = request.form.get('date')
            selected_period = request.form.get('selected_period')

            if not selected_period or not selected_period.isdigit():
                return "Error: No period selected or invalid value", 400

            for student in students:
                student_id = student['student_id']
                remarks = request.form.get(f'remarks_{student_id}', '')
                status_value = request.form.get(f'status_{student_id}', 'Absent')

                c2 = db.cursor(dictionary=True)
                try:
                    # check existing record
                    c2.execute(
                        "SELECT * FROM attendance WHERE student_id = %s AND date = %s",
                        (student_id, date)
                    )
                    existing = c2.fetchone()

                    if existing:
                        update_sql = f"""
                            UPDATE attendance 
                            SET period{selected_period} = %s, remarks = %s 
                            WHERE student_id = %s AND date = %s
                        """
                        c2.execute(update_sql, (status_value, remarks, student_id, date))
                        db.commit()
                    else:
                        # create new row with all 6 periods as NULL
                        period_fields = [None] * 6
                        period_fields[int(selected_period) - 1] = status_value
                        c2.execute(
                            """
                            INSERT INTO attendance 
                            (student_id, date, period1, period2, period3, period4, period5, period6, remarks, added_by) 
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            """,
                            (
                                student_id,
                                date,
                                period_fields[0],
                                period_fields[1],
                                period_fields[2],
                                period_fields[3],
                                period_fields[4],
                                period_fields[5],
                                remarks,
                                teacher_username
                            )
                        )
                        db.commit()
                finally:
                    c2.close()

            return redirect(url_for('view_attendance'))

        return render_template('mark_attendance.html', students=students, selected_period=selected_period)

    finally:
        cursor.close()


@app.route('/view_attendance', methods=['GET', 'POST'])
def view_attendance():
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor(dictionary=True)

    try:
        today = datetime.today().date()
        six_months_ago = today - timedelta(days=180)

        # default date = today
        selected_date = today

        if request.method == 'POST':
            sd = request.form.get('selected_date')
            if sd:
                try:
                    selected_date = datetime.strptime(sd, '%Y-%m-%d').date()
                except ValueError:
                    selected_date = today

        # ✅ Fetch only attendance of the selected date (not all)
        cursor.execute('''
            SELECT s.student_id, s.name, a.date,
                   a.period1, a.period2, a.period3, a.period4, a.period5, a.period6, a.remarks
            FROM attendance a
            JOIN students s ON a.student_id = s.student_id
            WHERE a.date = %s
            ORDER BY a.student_id
        ''', (selected_date,))
        records = cursor.fetchall()

        # ✅ Attendance summary only for selected day
        total_present = total_absent = total_slots = 0
        for row in records:
            for i in range(1, 7):  # check each period
                val = row.get(f'period{i}')
                if val:
                    total_slots += 1
                    if val.strip().lower() == 'present':
                        total_present += 1
                    elif val.strip().lower() == 'absent':
                        total_absent += 1

        attendance_percentage = round((total_present / total_slots) * 100, 2) if total_slots else None

        return render_template('view_attendance.html',
                               records=records,
                               total_present=total_present,
                               total_absent=total_absent,
                               attendance_percentage=attendance_percentage,
                               selected_date=selected_date)
    finally:
        cursor.close()


@app.route('/student_view_attendance')
def student_view_attendance():
    if 'user' not in session or session['role'] != 'student':
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        student_id = session['user']  # ✅ username == student_id

        # last 6 months filter
        six_months_ago_str = (datetime.now() - timedelta(days=180)).strftime('%Y-%m-%d')
        cursor.execute("""
            SELECT date, period1, period2, period3, period4, period5, period6, remarks
            FROM attendance
            WHERE student_id = %s AND date >= %s
            ORDER BY date DESC
        """, (student_id, six_months_ago_str))
        attendance = cursor.fetchall()

        # Attendance summary
        total_present = total_absent = total_periods = 0
        for rec in attendance:
            for i in range(1, 7):  # period1 .. period6
                status = rec.get(f'period{i}')
                if status:
                    total_periods += 1
                    if status.strip().lower() == 'present':
                        total_present += 1
                    elif status.strip().lower() == 'absent':
                        total_absent += 1

        attendance_percentage = (total_present / total_periods * 100) if total_periods > 0 else 0

        return render_template(
            'student_view_attendance.html',
            attendance=attendance,            # ✅ corrected (was "records")
            total_present=total_present,
            total_absent=total_absent,
            attendance_percentage=round(attendance_percentage, 2)
        )
    finally:
        cursor.close()

import io
import openpyxl
from openpyxl.styles import Font, Alignment

@app.route('/download_attendance_excel')
def download_attendance_excel():
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect('/')

    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute('''
            SELECT s.student_id, s.name, a.date, a.period1, a.period2, a.period3, a.period4, a.period5, a.period6
            FROM attendance a
            JOIN students s ON a.student_id = s.student_id
            ORDER BY a.date DESC
        ''')
        records = cursor.fetchall()

        summary = {}
        total_present = total_absent = 0
        for record in records:
            sid = record['student_id']
            name = record['name']
            summary.setdefault((sid, name), {'Present': 0, 'Absent': 0})
            for i in range(1, 7):
                status = record.get(f'period{i}')
                if status == 'Present':
                    summary[(sid, name)]['Present'] += 1
                    total_present += 1
                elif status == 'Absent':
                    summary[(sid, name)]['Absent'] += 1
                    total_absent += 1

        # ✅ Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Summary"

        # Header
        headers = ["Student ID", "Name", "Present Count", "Absent Count"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for (sid, name), counts in summary.items():
            ws.append([sid, name, counts['Present'], counts['Absent']])

        # Totals row
        ws.append([])
        ws.append(["", "TOTAL", total_present, total_absent])

        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        # Auto column width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # ✅ Save to BytesIO for download
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment;filename=attendance_summary.xlsx"}
        )
    finally:
        cursor.close()

# ---------------------
# Attendance percentage utilities
# ---------------------
@app.route('/attendance_percentage')
def attendance_percentage():
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect('/')

    db = get_db()
    cursor = db.cursor(dictionary=True)
    teacher_username = session['username']
    try:
        six_months_ago = (datetime.now() - timedelta(days=180)).strftime('%Y-%m-%d')
        cursor.execute("""
            SELECT s.username, s.name,
                   COUNT(a.id) AS total_days,
                   SUM(CASE WHEN a.period1 = 'Present' OR a.period2 = 'Present' OR a.period3 = 'Present' OR
                            a.period4 = 'Present' OR a.period5 = 'Present' OR a.period6 = 'Present' THEN 1 ELSE 0 END) AS present_days
            FROM students s
            LEFT JOIN attendance a ON s.username = a.username AND a.date >= %s
            WHERE s.added_by = %s
            GROUP BY s.username
        """, (six_months_ago, teacher_username))
        rows = cursor.fetchall()
        result = []
        for row in rows:
            total_days = row.get('total_days') or 0
            present_days = row.get('present_days') or 0
            percentage = round((present_days / total_days * 100), 2) if total_days else 0
            result.append({
                'student_id': row['username'],
                'name': row['name'],
                'percentage': percentage,
                'present_days': present_days,
                'total_days': total_days
            })
        return render_template('attendance_percentage.html', data=result)
    finally:
        cursor.close()

@app.route('/custom_attendance_percentage', methods=['GET', 'POST'])
def custom_attendance_percentage():
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect('/')

    db = get_db()
    teacher_username = session['username']
    cursor = db.cursor(dictionary=True)
    data = []
    start_date = end_date = None
    try:
        if request.method == 'POST':
            start_date = request.form.get('start_date')
            end_date = request.form.get('end_date')
            cursor.execute("""
                SELECT s.username, s.name,
                       COUNT(a.id) AS total_days,
                       SUM(CASE WHEN a.period1 = 'Present' OR a.period2 = 'Present' OR a.period3 = 'Present' OR
                                a.period4 = 'Present' OR a.period5 = 'Present' OR a.period6 = 'Present' THEN 1 ELSE 0 END) AS present_days
                FROM students s
                LEFT JOIN attendance a ON s.username = a.username AND a.date BETWEEN %s AND %s
                WHERE s.added_by = %s
                GROUP BY s.username
            """, (start_date, end_date, teacher_username))
            rows = cursor.fetchall()
            for row in rows:
                total_days = row.get('total_days') or 0
                present_days = row.get('present_days') or 0
                percentage = round((present_days / total_days * 100), 2) if total_days else 0
                data.append({
                    'student_id': row['username'],
                    'name': row['name'],
                    'percentage': percentage,
                    'present_days': present_days,
                    'total_days': total_days
                })
        return render_template('custom_attendance_percentage.html', data=data, start_date=start_date, end_date=end_date)
    finally:
        cursor.close()

# ---------------------
# Password reset (forgot/verify/reset)
# ---------------------
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    error = None
    if request.method == 'POST':
        username = request.form.get('username')
        role = request.form.get('role')
        db = get_db()
        cursor = db.cursor(dictionary=True)
        try:
            if role == 'student':
                cursor.execute("SELECT email FROM students WHERE student_id = %s", (username,))
            else:
                cursor.execute("SELECT email FROM users WHERE username = %s AND role = 'teacher'", (username,))
            user = cursor.fetchone()
            if user and user.get('email'):
                email = user['email']
                otp = str(random.randint(100000, 999999))
                session['otp'] = otp
                session['reset_username'] = username
                session['reset_role'] = role
                msg = Message('Your OTP Code', recipients=[email])
                msg.body = f'Your OTP is: {otp}'
                mail.send(msg)
                return redirect(url_for('verify_otp'))
            else:
                error = "No account found with that username."
        finally:
            cursor.close()
    return render_template('forgot_password.html', error=error)

@app.route('/verify_otp', methods=['GET', 'POST'])
def verify_otp():
    error = None
    if request.method == 'POST':
        entered_otp = request.form.get('otp')
        if entered_otp == session.get('otp'):
            return redirect(url_for('reset_password'))
        else:
            error = "Invalid OTP. Please try again."
    return render_template('verify_otp.html', error=error)

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    error = None
    success = None
    if request.method == 'POST':
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        if new_password != confirm_password:
            error = "Passwords do not match."
        elif len(new_password) < 6:
            error = "Password must be at least 6 characters long."
        else:
            hashed_pw = generate_password_hash(new_password)
            username = session.get('reset_username')
            role = session.get('reset_role')
            if not username or not role:
                error = "Session expired or invalid access."
                return render_template('reset_password.html', error=error)
            db = get_db()
            cursor = db.cursor()
            try:
                if role == 'student':
                    # update users table password for the student's username if mapped, or update students table if you store password there
                    # We'll update users table where username matches student's username if such exists
                    cursor.execute("UPDATE users SET password = %s WHERE username = %s AND role = 'student'",
                                   (hashed_pw, username))
                    db.commit()
                    return redirect(url_for('student_login'))
                elif role == 'teacher':
                    cursor.execute("UPDATE users SET password = %s WHERE username = %s AND role = 'teacher'",
                                   (hashed_pw, username))
                    db.commit()
                    return redirect(url_for('teacher_login'))
                else:
                    error = "Invalid user role."
            except Exception as e:
                db.rollback()
                error = f"Error resetting password: {e}"
            finally:
                cursor.close()
    return render_template('reset_password.html', error=error, success=success)

# ---------------------
# Internal marks
# ---------------------
# ---------------- INTERNAL MARKS ROUTES ---------------- #

@app.route('/internal_marks', methods=['GET', 'POST'])
def internal_marks():
    # ✅ Only teachers can access
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM students")
        students = cursor.fetchall()
        error = None

        if request.method == 'POST':
            student_id = request.form.get('student_id')
            semester = request.form.get('semester')

            try:
                c2 = db.cursor()
                try:
                    for i in range(1, 6):  # 5 subjects per upload
                        subject = request.form.get(f'subject_{i}')
                        marks = request.form.get(f'marks_{i}')
                        if subject and marks:
                            c2.execute(
                                "INSERT INTO internal_marks (student_id, subject, marks, semester, entered_by) "
                                "VALUES (%s, %s, %s, %s, %s)",
                                (student_id, subject, marks, semester, session['username'])
                            )
                    db.commit()
                    return redirect(url_for('view_internal_marks'))
                finally:
                    c2.close()
            except mysql.connector.IntegrityError as e:
                db.rollback()
                error = "Duplicate subject/semester entry or error: " + str(e)

        return render_template('internal_marks.html', students=students, error=error)
    finally:
        cursor.close()


@app.route('/enter_internal_marks', methods=['GET', 'POST'])
def enter_internal_marks():
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute("SELECT id, name, student_id FROM students")
        students = cursor.fetchall()

        if request.method == 'POST':
            student_id = request.form.get('student_id')
            subject = request.form.get('subject')
            marks = request.form.get('marks')
            semester = request.form.get('semester')

            c2 = db.cursor()
            try:
                c2.execute(
                    "INSERT INTO internal_marks (student_id, subject, marks, semester, entered_by) "
                    "VALUES (%s,%s,%s,%s,%s)",
                    (student_id, subject, marks, semester, session['username'])
                )
                db.commit()
                return redirect(url_for('view_internal_marks'))
            finally:
                c2.close()

        return render_template('enter_internal_marks.html', students=students)
    finally:
        cursor.close()



@app.route('/view_internal_marks', methods=['GET', 'POST'])
def view_internal_marks():
    if 'username' not in session and 'user' not in session:
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor(dictionary=True)
    role = session.get('role')

    try:
        # ----------------- Student View -----------------
        if role == 'student':
            student_id = session['user']  # student_id stored in session
            semester = request.args.get('semester')  # dropdown filter

            query = '''
                SELECT im.student_id, s.name AS student_name, im.subject, im.marks, im.semester
                FROM internal_marks im
                JOIN students s ON im.student_id = s.student_id
                WHERE im.student_id = %s
            '''
            params = [student_id]

            if semester and semester.isdigit():
                query += " AND im.semester = %s"
                params.append(int(semester))

            cursor.execute(query, tuple(params))
            records = cursor.fetchall()

            return render_template(
                'view_internal_marks.html',
                records=records,
                students=None,   # student ki dropdown avasaram ledu
                selected_semester=semester
            )

        # ----------------- Teacher View -----------------
        elif role == 'teacher':
            teacher_username = session['username']

            # Filters from form
            student_id = request.form.get('student_id') if request.method == 'POST' else None
            subject = request.form.get('subject') if request.method == 'POST' else None
            semester = request.form.get('semester') if request.method == 'POST' else None

            query = '''
                SELECT im.student_id, s.name AS student_name, im.subject, im.marks, im.semester
                FROM internal_marks im
                JOIN students s ON im.student_id = s.student_id
            '''
            params = []
            conditions = []

            if student_id:
                conditions.append('im.student_id = %s')
                params.append(student_id)
            if subject:
                conditions.append('im.subject LIKE %s')
                params.append(f'%{subject}%')
            if semester and semester.isdigit():
                conditions.append('im.semester = %s')
                params.append(int(semester))

            if conditions:
                query += ' WHERE ' + ' AND '.join(conditions)

            cursor.execute(query, tuple(params))
            records = cursor.fetchall()

            # Dropdown list of all students
            cursor.execute("SELECT student_id, name FROM students")
            students = cursor.fetchall()

            return render_template(
                'view_internal_marks.html',
                records=records,
                students=students,
                selected_semester=semester
            )

        else:
            return redirect(url_for('choose_login'))

    finally:
        cursor.close()



@app.route('/delete_internal_marks/<student_id>/<subject>/<semester>')
def delete_internal_marks(student_id, subject, semester):
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute(
            "DELETE FROM internal_marks WHERE student_id = %s AND subject = %s AND semester = %s",
            (student_id, subject, semester)
        )
        db.commit()
    finally:
        cursor.close()
    return redirect(url_for('view_internal_marks'))


@app.route('/edit_internal_marks/<student_id>/<subject>/<semester>', methods=['GET', 'POST'])
def edit_internal_marks(student_id, subject, semester):
    if 'username' not in session or session.get('role') != 'teacher':
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        if request.method == 'POST':
            new_marks = request.form.get('marks')
            cur2 = db.cursor()
            try:
                cur2.execute(
                    "UPDATE internal_marks SET marks = %s WHERE student_id = %s AND subject = %s AND semester = %s",
                    (new_marks, student_id, subject, semester)
                )
                db.commit()
                return redirect(url_for('view_internal_marks'))
            finally:
                cur2.close()

        cursor.execute(
            "SELECT * FROM internal_marks WHERE student_id = %s AND subject = %s AND semester = %s",
            (student_id, subject, semester)
        )
        row = cursor.fetchone()
        marks = row['marks'] if row else None

        return render_template('edit_internal_marks.html',
                               student_id=student_id,
                               subject=subject,
                               semester=semester,
                               marks=marks)
    finally:
        cursor.close()


# ---------------------
# Messaging (send/inbox)
# ---------------------
@app.route('/send_message', methods=['GET', 'POST'])
def send_message():
    if 'user' not in session:
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        role = session.get('role')
        sender = session['user']

        if role == 'teacher':
            cursor.execute("SELECT username FROM users WHERE role = 'student'")
        elif role == 'student':
            cursor.execute("SELECT username FROM users WHERE role = 'teacher'")
        else:
            return redirect(url_for('choose_login'))

        users = cursor.fetchall()

        if request.method == 'POST':
            receiver = request.form.get('receiver')
            message = request.form.get('message', '').strip()

            if not message:
                return render_template('send_message.html', users=users, error="Message cannot be empty")

            cur2 = db.cursor()
            try:
                cur2.execute(
    "INSERT INTO messages (sender, receiver, message, sender_role) VALUES (%s, %s, %s, %s)",
    (sender, receiver, message, role)
)

                db.commit()
                return redirect(url_for('inbox'))
            finally:
                cur2.close()

        return render_template('send_message.html', users=users)
    finally:
        cursor.close()



@app.route('/inbox')
def inbox():
    if 'user' not in session:
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor(dictionary=True)

    username = session['user']
    role = session.get('role')

    if role == 'teacher':
        # Show only messages sent by students
        cursor.execute("SELECT * FROM messages WHERE receiver = %s AND sender_role = 'student' ORDER BY timestamp DESC", (username,))
    elif role == 'student':
        # Show only messages sent by teachers
        cursor.execute("SELECT * FROM messages WHERE receiver = %s AND sender_role = 'teacher' ORDER BY timestamp DESC", (username,))
    else:
        return redirect(url_for('choose_login'))

    messages = cursor.fetchall()
    cursor.close()
    return render_template('inbox.html', messages=messages)

import csv
import io
from werkzeug.utils import secure_filename

# Allowed extensions
ALLOWED_EXTENSIONS = {'csv'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


import os, io, csv
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash

ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls', 'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


import os, io, csv
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash
import openpyxl  # for Excel files

ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls', 'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


import os, io, csv
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash
import openpyxl  # for Excel files
import pdfplumber  # for PDF parsing

ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls', 'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/upload/<datatype>', methods=['GET', 'POST'])
def upload_file(datatype):
    """
    datatype = 'students', 'marks', 'attendance'
    Upload a CSV, Excel, or PDF file.
    """

    print("DEBUG session contents:", dict(session))

    # --- Check login & role ---
    user = session.get('username')
    role = session.get('role')

    if not user:
        flash("⚠️ Please log in first.", "error")
        return redirect(url_for('choose_login'))

    if role != 'teacher':
        flash("⛔ Only teachers can upload files.", "error")
        return redirect(url_for('teacher_dashboard'))

    message, message_type = None, None

    if request.method == 'POST':
        if 'file' not in request.files:
            return render_template('upload.html', message="⚠️ No file uploaded.", message_type="error", datatype=datatype)

        file = request.files['file']

        if file.filename == '':
            return render_template('upload.html', message="⚠️ No file selected.", message_type="error", datatype=datatype)

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            ext = filename.rsplit('.', 1)[1].lower()

            db = get_db()
            cursor = db.cursor()

            try:
                # ----------------- CSV -----------------
                if ext == "csv":
                    stream = io.StringIO(file.stream.read().decode("utf-8"), newline=None)
                    csv_input = csv.reader(stream)
                    header = next(csv_input, None)

                    if datatype == "students":
                        for row in csv_input:
                            if len(row) < 3:
                                continue
                            raw_username, name, branch = [col.strip() for col in row[:3]]
                            student_id = raw_username.replace("_", "")
                            email = f"{student_id.lower()}@college.com"

                            # Insert into users
                            cursor.execute(
                                """
                                INSERT INTO users (username, email, role)
                                VALUES (%s, %s, 'student')
                                ON DUPLICATE KEY UPDATE email = VALUES(email), role = VALUES(role)
                                """,
                                (student_id, email)
                            )

                            # Insert into students
                            cursor.execute(
                                """
                                INSERT INTO students (student_id, username, name, email, course)
                                VALUES (%s, %s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE 
                                    name = VALUES(name), 
                                    email = VALUES(email), 
                                    course = VALUES(course)
                                """,
                                (student_id, student_id, name, email, branch)
                            )
                        db.commit()

                    elif datatype == "marks":
                        for row in csv_input:
                            if len(row) < 3:
                                continue
                            student_id, subject, marks = [col.strip() for col in row[:3]]
                            student_id = student_id.replace("_", "")
                            cursor.execute(
                                "INSERT INTO internal_marks (student_id, subject, marks, entered_by) VALUES (%s, %s, %s, %s)",
                                (student_id, subject, marks, user)
                            )
                        db.commit()

                    elif datatype == "attendance":
                        for row in csv_input:
                            if len(row) < 3:
                                continue
                            student_id, date, status = [col.strip() for col in row[:3]]
                            student_id = student_id.replace("_", "")
                            cursor.execute(
                                """
                                INSERT INTO attendance (student_id, username, date, status, added_by)
                                VALUES (%s, %s, %s, %s, %s)
                                """,
                                (student_id, student_id, date, status, user)
                            )
                        db.commit()

                # ----------------- Excel -----------------
                elif ext in ["xlsx", "xls"]:
                    wb = openpyxl.load_workbook(file)
                    sheet = wb.active
                    rows = list(sheet.iter_rows(values_only=True))
                    header = rows[0] if rows else None

                    if datatype == "students":
                        for row in rows[1:]:
                            if len(row) < 3:
                                continue
                            raw_username, name, branch = [str(col).strip() for col in row[:3]]
                            student_id = raw_username.replace("_", "")
                            email = f"{student_id.lower()}@gmail.com"
                            default_pw = generate_password_hash(student_id)

                            cursor.execute(
                                """
                                INSERT INTO users (username, email, role,password)
                                VALUES (%s, %s, 'student',%s)
                                ON DUPLICATE KEY UPDATE email = VALUES(email), role = VALUES(role)
                                """,
                                (student_id, email,default_pw)
                            )

                            cursor.execute(
                                """
                                INSERT INTO students (student_id, username, name, email, course)
                                VALUES (%s, %s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE 
                                    name = VALUES(name), 
                                    email = VALUES(email), 
                                    course = VALUES(course)
                                """,
                                (student_id, student_id, name, email, branch)
                            )
                        db.commit()

                    elif datatype == "marks":
                        for row in rows[1:]:
                            if len(row) < 3:
                                continue
                            student_id, subject, marks = [str(col).strip() for col in row[:3]]
                            student_id = student_id.replace("_", "")
                            cursor.execute(
                                "INSERT INTO internal_marks (student_id, subject, marks, entered_by) VALUES (%s, %s, %s, %s)",
                                (student_id, subject, marks, user)
                            )
                        db.commit()

                    elif datatype == "attendance":
                        for row in rows[1:]:
                            if len(row) < 3:
                                continue
                            student_id, date, status = [str(col).strip() for col in row[:3]]
                            student_id = student_id.replace("_", "")
                            cursor.execute(
                                """
                                INSERT INTO attendance (student_id, username, date, status, added_by)
                                VALUES (%s, %s, %s, %s, %s)
                                """,
                                (student_id, student_id, date, status, user)
                            )
                        db.commit()

                # ----------------- PDF -----------------
                elif ext == "pdf":
                    filepath = os.path.join("uploads", filename)
                    os.makedirs("uploads", exist_ok=True)
                    file.save(filepath)

                    with pdfplumber.open(filepath) as pdf:
                        for page in pdf.pages:
                            table = page.extract_table()
                            if not table:
                                continue
                            for row in table[1:]:  # skip header
                                if datatype == "students":
                                    raw_username, name, branch = [col.strip() for col in row[:3]]
                                    student_id = raw_username.replace("_", "")
                                    email = f"{student_id.lower()}@gmail.com"

                                    cursor.execute(
                                        """
                                        INSERT INTO users (username, email, role)
                                        VALUES (%s, %s, 'student')
                                        ON DUPLICATE KEY UPDATE email = VALUES(email), role = VALUES(role)
                                        """,
                                        (student_id, email)
                                    )

                                    cursor.execute(
                                        """
                                        INSERT INTO students (student_id, username, name, email, course)
                                        VALUES (%s, %s, %s, %s, %s)
                                        ON DUPLICATE KEY UPDATE 
                                            name = VALUES(name), 
                                            email = VALUES(email), 
                                            course = VALUES(course)
                                        """,
                                        (student_id, student_id, name, email, branch)
                                    )

                                elif datatype == "marks":
                                    student_id, subject, marks = [col.strip() for col in row[:3]]
                                    student_id = student_id.replace("_", "")
                                    cursor.execute(
                                        "INSERT INTO internal_marks (student_id, subject, marks, entered_by) VALUES (%s, %s, %s, %s)",
                                        (student_id, subject, marks, user)
                                    )

                                elif datatype == "attendance":
                                    student_id, date, status = [col.strip() for col in row[:3]]
                                    student_id = student_id.replace("_", "")
                                    cursor.execute(
                                        """
                                        INSERT INTO attendance (student_id, username, date, status, added_by)
                                        VALUES (%s, %s, %s, %s, %s)
                                        """,
                                        (student_id, student_id, date, status, user)
                                    )
                    db.commit()

                # ----------------- Log upload -----------------
                cursor.execute(
                    """INSERT INTO faculty_uploads (file_name, file_type, uploaded_by, datatype, uploaded_at) 
                       VALUES (%s, %s, %s, %s, NOW())""",
                    (filename, ext, user, datatype)
                )
                db.commit()

                message, message_type = f"✅ File '{filename}' uploaded successfully!", "success"

            except Exception as e:
                db.rollback()
                message, message_type = f"❌ Error uploading file: {str(e)}", "error"
            finally:
                cursor.close()

        else:
            message, message_type = "❌ Invalid file type. Allowed: CSV, Excel, PDF", "error"

    return render_template("upload.html", message=message, message_type=message_type, datatype=datatype)

@app.route('/debug_data/<table>')
def debug_data(table):
    """
    Temporary route: Show contents of a table after upload.
    Example: /debug_data/students or /debug_data/attendance
    """
    if 'user' not in session or session.get('role') != 'teacher':
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor(dictionary=True)

    try:
        cursor.execute(f"SELECT * FROM {table} LIMIT 50")  # avoid huge dumps
        rows = cursor.fetchall()
        return {"data": rows}  # JSON output for quick checking
    except Exception as e:
        return {"error": str(e)}

@app.route('/routedebug/<table>', endpoint='routedebug_data')
def debug_data(table):
    if 'user' not in session or session.get('role') != 'teacher':
        return redirect(url_for('choose_login'))

    db = get_db()
    cursor = db.cursor(dictionary=True)

    try:
        cursor.execute(f"SELECT * FROM {table}")
        rows = cursor.fetchall()
        return render_template("debug.html", table=table, rows=rows)
    except Exception as e:
        return f"Error: {e}"
    finally:
        cursor.close()
        db.close()
from flask import (
    Flask, render_template, request, redirect,
    url_for, flash, session
)
from werkzeug.utils import secure_filename
import os
from datetime import datetime

# --- Upload Notes ---
@app.route('/upload_notes', methods=['GET', 'POST'])
def upload_notes():
    if 'role' not in session or session['role'] != 'teacher':
        flash("⚠ Please login as a teacher first.", "danger")
        return redirect(url_for('choose_login'))

    if request.method == 'POST':
        subject = request.form.get('subject', '').strip()
        department = request.form.get('department', '').strip()
        year = request.form.get('year', '').strip()
        title = request.form.get('title', '').strip()
        file = request.files.get('file')

        # --- Validation ---
        if not subject or not department or not year or not title or not file:
            flash("❌ All fields are required.", "danger")
            return redirect(request.url)

        try:
            year = int(year)
        except ValueError:
            flash("❌ Year must be a valid number.", "danger")
            return redirect(request.url)

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

            try:
                # Save file
                file.save(filepath)

                # Save DB record
                db = get_db()
                cur = db.cursor(dictionary=True)
                cur.execute("""
                    INSERT INTO notes 
                        (subject, department, year, title, filename, uploaded_by, uploaded_on) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (subject, department, year, title, filename, session['username'], datetime.now()))
                db.commit()

                flash(f"✅ '{title}' uploaded successfully!", "success")

            except Exception as e:
                db.rollback()
                flash(f"❌ Error uploading notes: {e}", "danger")

            finally:
                cur.close()

            return redirect(url_for('upload_notes'))

        else:
            flash("❌ Invalid file type. Allowed: pdf, docx, pptx, txt", "danger")
            return redirect(request.url)

    return render_template('upload_notes.html')


# --- View Notes (for Students) ---
@app.route('/view_notes')
def view_notes():
    if 'role' not in session or session['role'] != 'student':
        flash("⚠ Please login as a student first.", "danger")
        return redirect(url_for('choose_login'))

    dept = request.args.get('department', '').strip()
    year = request.args.get('year', '').strip()

    db = get_db()
    cur = db.cursor(dictionary=True)
    notes = []

    try:
        query = "SELECT * FROM notes WHERE 1=1"
        params = []

        if dept:
            query += " AND department = %s"
            params.append(dept)
        if year:
            try:
                year = int(year)
                query += " AND year = %s"
                params.append(year)
            except ValueError:
                flash("❌ Year filter must be a number.", "danger")

        query += " ORDER BY uploaded_on DESC"
        cur.execute(query, params)
        notes = cur.fetchall()

    except Exception as e:
        flash(f"❌ Error fetching notes: {e}", "danger")

    finally:
        cur.close()

    return render_template('view_notes.html', notes=notes)

# Root
# ---------------------
@app.route('/')
def choose_login():
    return render_template('choose_login.html')

# ---------------------
# Run
# ---------------------
if __name__ == '__main__':
    init_db()
    app.run(debug=True)
